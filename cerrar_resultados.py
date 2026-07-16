"""
cerrar_resultados.py
---------------------
FASE 4, versión 3. Reintenta cada 15 min entre las 23:00 y la 01:00 (ver
el workflow) hasta lograrlo, en vez de depender de un solo disparo exacto
a las 23:30 — esto fue justo lo que falló: el día que probamos, esa
ejecución puntual no llegó a correr a tiempo (o no corrió), y como no
había ningún reintento, ningún partido quedó resuelto ese día.

Hace 3 cosas:
  1. Consulta el resultado final de cada partido seleccionado hoy y
     calcula si acertó el favorito (ganó) o no.
  2. Archiva el día completo en data/historial_dias/{fecha}.json.
  3. Actualiza data/estadisticas.xlsx (2 pestañas).
"""

import json
from pathlib import Path

from fetch_data import obtener_resultado_fixture
from cuota_api_football import uso_de_hoy
from estado_diario import ya_se_hizo, marcar_hecho

DATA_DIR = Path(__file__).parent / "data"
ARCHIVO_PARTIDOS = DATA_DIR / "partidos_hoy.json"
DIR_HISTORIAL_DIAS = DATA_DIR / "historial_dias"
ARCHIVO_EXCEL = DATA_DIR / "estadisticas.xlsx"

ESTADOS_TERMINADO = ("FT", "AET", "PEN")


def _calcular_acierto(p, goles_local, goles_visitante):
    goles_favorito = goles_local if p["favorito_es_local"] else goles_visitante
    goles_rival = goles_visitante if p["favorito_es_local"] else goles_local
    return goles_favorito > goles_rival


def cerrar():
    if ya_se_hizo("cierre"):
        print("El cierre de hoy ya se hizo antes. Nada que hacer.")
        return

    if not ARCHIVO_PARTIDOS.exists():
        print("No hay partidos_hoy.json todavía. Se reintentará en el próximo ciclo.")
        return

    datos = json.loads(ARCHIVO_PARTIDOS.read_text(encoding="utf-8"))
    cambios = False

    for p in datos["partidos"]:
        if p.get("acierto") is not None or not p.get("fixture_id"):
            continue
        try:
            info = obtener_resultado_fixture(p["fixture_id"])
        except Exception as e:
            print(f"[AVISO] No se pudo consultar el resultado de {p['partido']}: {e}")
            continue
        if not info:
            continue

        estado = info["fixture"]["status"]["short"]
        if estado not in ESTADOS_TERMINADO:
            continue

        gh, ga = info["goals"]["home"], info["goals"]["away"]
        p["resultado_final"] = f"{gh}-{ga}"
        p["acierto"] = _calcular_acierto(p, gh, ga)
        cambios = True

    if cambios:
        ARCHIVO_PARTIDOS.write_text(json.dumps(datos, ensure_ascii=False, indent=2), encoding="utf-8")

    usadas, disponibles = uso_de_hoy()

    DIR_HISTORIAL_DIAS.mkdir(exist_ok=True, parents=True)
    archivo_dia = DIR_HISTORIAL_DIAS / f"{datos['fecha']}.json"
    archivo_dia.write_text(json.dumps({
        "fecha": datos["fecha"],
        "partidos": datos["partidos"],
        "api_football_usadas": usadas,
        "api_football_disponibles": disponibles,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Día archivado en {archivo_dia}")

    _actualizar_excel(datos["fecha"], datos["partidos"], usadas, disponibles)
    marcar_hecho("cierre")


def _actualizar_excel(fecha, partidos, usadas, disponibles):
    try:
        import openpyxl
    except ImportError:
        print("[AVISO] openpyxl no está instalado, no se pudo actualizar el Excel "
              "(agrega 'openpyxl' a requirements.txt).")
        return

    if ARCHIVO_EXCEL.exists():
        wb = openpyxl.load_workbook(ARCHIVO_EXCEL)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        hoja1 = wb.create_sheet("Resultados diarios")
        hoja1.append(["Fecha", "Partido", "Favorito", "Local/Visitante", "Cuota inicial",
                      "Prob. inicial %", "Marcador final", "Acierto", "Alertas enviadas"])
        hoja2 = wb.create_sheet("Resumen por dia")
        hoja2.append(["Fecha", "Total partidos", "Aciertos", "% Aciertos",
                      "Peticiones API usadas", "Peticiones disponibles"])

    hoja1 = wb["Resultados diarios"]
    hoja2 = wb["Resumen por dia"]

    ya_registrado = any(fila[0].value == fecha for fila in hoja2.iter_rows(min_row=2) if fila[0].value)
    if ya_registrado:
        print(f"El día {fecha} ya estaba registrado en el Excel, no se duplica.")
        return

    total = len(partidos)
    aciertos = sum(1 for p in partidos if p.get("acierto") is True)
    resueltos = sum(1 for p in partidos if p.get("acierto") is not None)

    for p in partidos:
        hoja1.append([
            fecha, p["partido"], p["favorito"],
            "Local" if p["favorito_es_local"] else "Visitante",
            p["cuota_inicial"], p["probabilidad_inicial"],
            p.get("resultado_final") or "sin resolver",
            "✅" if p.get("acierto") is True else ("❌" if p.get("acierto") is False else "?"),
            len(p.get("alertas_enviadas", [])),
        ])

    pct_aciertos = round((aciertos / resueltos) * 100, 1) if resueltos else None
    hoja2.append([fecha, total, aciertos, pct_aciertos, usadas, disponibles])

    DATA_DIR.mkdir(exist_ok=True)
    wb.save(ARCHIVO_EXCEL)
    print(f"Excel actualizado: {ARCHIVO_EXCEL} ({total} partidos, {aciertos} aciertos)")


if __name__ == "__main__":
    cerrar()
